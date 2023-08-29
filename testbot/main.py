# Подключаем библиотеку
import telebot
from pydub import AudioSegment


import os
import speech_recognition
from pydub import AudioSegment
from openpyxl import Workbook
import requests
from bs4 import BeautifulSoup


# Здесь нужно вставить токен, который дал BotFather при регистрации
# Пример: token = '2007628239:AAEF4ZVqLiRKG7j49EC4vaRwXjJ6DN6xng8'
token = '6284689879:AAEwDZh4aL5i1e5cn_HB4BVEwuBwVfejBiU'  # <<< Ваш токен

# В этой строчке мы создаем бота и даем ему запомнить токен
#из библиотеки telebot через . получаем нужный обьект
bot = telebot.TeleBot(token)
'''
# Пишем первую функцию, которая отвечает "Привет" на команду /start
# Все функции общения приложения с ТГ спрятаны в функции под обертку @
@bot.message_handler(commands=['start'])
def say_hi(message):
    bot.send_message(message.chat.id, 'Привет')
#send_message-функция отправить сообщение т

# Запускаем бота. Он будет работать до тех пор, пока работает ячейка (крутится значок слева).
# Остановим ячейку - остановится бот
bot.polling() #запуск
'''




#функция преобразования голоса

def oga2wav(filename):
    # Переименование формата: 'sample.oga' -> 'sample.wav'
    new_filename = filename.replace('.oga','.wav')
    # Читаем файл с диска с помощью функции AudioSegment.from_file()
    audio = AudioSegment.from_file(filename)
    # Экспортируем файл в новом формате
    audio.export(new_filename, format='wav')
    # Возвратим в качестве результата функции имя нового файла
    return new_filename

def recognize_speech(oga_filename):
    # Перевод голоса в текст + удаление использованных файлов
    wav_filename = oga2wav(oga_filename)
    recognizer = speech_recognition.Recognizer()

    with speech_recognition.WavFile(wav_filename) as source:
        wav_audio = recognizer.record(source)

    text = recognizer.recognize_google(wav_audio, language='ru')

    if os.path.exists(oga_filename):
        os.remove(oga_filename)

    if os.path.exists(wav_filename):
        os.remove(wav_filename)

    return text







def download_file(bot, file_id):
    # Скачивание файла, который прислал пользователь
    file_info = bot.get_file(file_id)
    downloaded_file = bot.download_file(file_info.file_path)
    filename = file_id + file_info.file_path
    filename = filename.replace('/', '_')
    with open(filename, 'wb') as f:
        f.write(downloaded_file)
    return filename


@bot.message_handler(commands=['start'])
def say_hi(message):
    # Функция, отправляющая "Привет" в ответ на команду /start
    bot.send_message(message.chat.id, 'Привет, '+message.chat.first_name+" "+message.chat.last_name)


@bot.message_handler(content_types=['voice'])
def transcript(message):
    # Функция, отправляющая текст в ответ на голосовое
    filename = download_file(bot, message.voice.file_id)
    text = recognize_speech(filename)
    bot.send_message(message.chat.id, text)

'''
# ↓↓↓ Пусть функция реагирует на изображения
@bot.message_handler(content_types=['photo'])
def resend_photo(message):
    # Скачиваем последний файл в списке с максимальным разрешением по file_id
    file_id = message.photo[-1].file_id
    filename = download_file(bot, file_id)

    # Открываем изображение из файла с помощью функции open, 'rb' = read bytes
    image = open(filename, 'rb')

    # Отправляем изображение в чат с пользователем
    bot.send_photo(message.chat.id, image)

    # Не забываем закрыть файл
    image.close()
'''


# Запускаем бота. Он будет работать до тех пор, пока работает ячейка (крутится значок слева).
# Остановим ячейку - остановится бот
bot.polling()



#РАБОТА С EXCEL
table = [[1, 2, 3], [4, 5, 6]]
# Создаем объект "книги" EXCEL
work_book = Workbook()

# Создаем объект "листа" EXCEL
work_sheet = work_book.active

# Для каждой строки row в списке table: присоединить строку к листу
for row in table:
    work_sheet.append(row)

work_book.save('table.xlsx')

work_book = Workbook()
work_sheet = work_book.active

# Коллекция для генерации таблицы
numbers = [1, 2, 3, 4, 5]

for elem in numbers:
    row = [elem, elem * 2, elem ** 2]
    print(row)
    work_sheet.append(row)

work_book.save('numbers.xlsx')



web_page = requests.get('https://live.skillbox.ru/playlists/code/python/')

web_page.text

soup = BeautifulSoup(web_page.text, 'html.parser')

soup.find(class_='playlist-inner-card__title hover-card__text t t--3').text

relative_url = soup.find(class_='playlist-inner-card hover-card').attrs['href']

abs_url = 'https://live.skillbox.ru'+relative_url

work_book = Workbook()
work_sheet = work_book.active

items = soup.find_all(class_='playlist-inner__item')

for elem in items:
    title = elem.find(class_='playlist-inner-card__title hover-card__text t t--3').text
    relative_url = elem.find(class_='playlist-inner-card hover-card').attrs['href']
    url = 'https://live.skillbox.ru' + relative_url
    row = [title, url]
    print(row)
    work_sheet.append(row)

work_book.save('Вебинары про Python от Skillbox.xlsx')
